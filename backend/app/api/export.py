"""Endpoints for data export, schema management, and backup management."""

from fastapi import APIRouter, Depends, HTTPException, Query, status
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import text
import logging
import json
import hashlib
from datetime import datetime
from typing import Optional
from io import BytesIO

from app.database.orm import get_db
from app.security import get_current_user
from app.models import EsquemaBaseDatos
from app.utils.exports import (
    export_table_to_csv,
    export_datos_importados_to_csv,
    export_to_excel,
    export_schema_to_json,
)
from app.utils.backup_manager import BackupManager

logger = logging.getLogger(__name__)
router = APIRouter(prefix="/api/export", tags=["Export & Schema"])


# ===== EXPORTACIÓN =====

@router.get("/table/{db_name}/{table_name}")
async def export_table_csv(
    db_name: str,
    table_name: str,
    format: str = Query("csv", regex="^(csv|excel)$"),
    limit: Optional[int] = Query(None),
    current_user: dict = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Export table data to CSV or Excel."""
    try:
        logger.info(f"{current_user['username']} exporting {db_name}.{table_name} to {format}")
        
        if format == "csv":
            csv_data = export_table_to_csv(db, db_name, table_name, limit)
            return StreamingResponse(
                iter([csv_data]),
                media_type="text/csv",
                headers={"Content-Disposition": f"attachment; filename={table_name}.csv"}
            )
        else:  # excel
            excel_bytes = export_to_excel(db, db_name, table_name, limit)
            return StreamingResponse(
                iter([excel_bytes]),
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f"attachment; filename={table_name}.xlsx"}
            )
    
    except Exception as e:
        logger.error(f"Export error: {e}")
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(e))


@router.get("/agentes")
async def export_agentes(
    format: str = Query("csv", regex="^(csv|excel)$"),
    with_pagos: bool = Query(False),
    current_user: dict = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Export agents/datos_importados with optional payment data."""
    try:
        logger.info(f"{current_user['username']} exporting agentes to {format}")
        
        csv_data = export_datos_importados_to_csv(db, with_pagos=with_pagos)
        
        if format == "csv":
            return StreamingResponse(
                iter([csv_data]),
                media_type="text/csv",
                headers={"Content-Disposition": "attachment; filename=agentes.csv"}
            )
        else:  # excel
            # Convert CSV string to Excel
            import csv
            from io import StringIO
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            
            reader = csv.reader(StringIO(csv_data))
            rows = list(reader)
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Agentes"
            
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            for row_idx, row in enumerate(rows, 1):
                for col_idx, cell_value in enumerate(row, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=cell_value)
                    if row_idx == 1:
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal="center")
            
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            
            return StreamingResponse(
                iter([output.getvalue()]),
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": "attachment; filename=agentes.xlsx"}
            )
    
    except Exception as e:
        logger.error(f"Export agentes error: {e}")
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(e))


# ===== GESTIÓN DE ESQUEMAS =====

@router.get("/schemas/{db_name}")
async def export_database_schema(
    db_name: str,
    current_user: dict = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Export complete database schema as JSON."""
    try:
        logger.info(f"{current_user['username']} exporting schema for {db_name}")
        
        schema_json = export_schema_to_json(db, db_name)
        
        return StreamingResponse(
            iter([schema_json]),
            media_type="application/json",
            headers={"Content-Disposition": f"attachment; filename={db_name}_schema.json"}
        )
    
    except Exception as e:
        logger.error(f"Schema export error: {e}")
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(e))


@router.post("/schemas/{db_name}/save")
async def save_schema_version(
    db_name: str,
    payload: dict,
    current_user: dict = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Save a versioned schema snapshot."""
    try:
        version = str((payload or {}).get("version", "1.0.0")).strip()
        etiqueta = str((payload or {}).get("etiqueta", "")).strip() or None
        descripcion = str((payload or {}).get("descripcion", "")).strip() or None
        
        logger.info(f"{current_user['username']} saving schema version {version} for {db_name}")
        
        # Get current schema
        schema_json = export_schema_to_json(db, db_name)
        hash_value = hashlib.sha256(schema_json.encode()).hexdigest()
        
        # Check for changes from previous version
        previous = db.query(EsquemaBaseDatos).filter(
            EsquemaBaseDatos.nombre_bd == db_name,
            EsquemaBaseDatos.activo == True,
        ).order_by(EsquemaBaseDatos.fecha_guardado.desc()).first()
        
        cambios = None
        if previous:
            prev_data = json.loads(previous.esquema_json)
            curr_data = json.loads(schema_json)
            cambios = _compare_schemas(prev_data, curr_data)
        
        # Save new version
        esquema = EsquemaBaseDatos(
            nombre_bd=db_name,
            version=version,
            etiqueta=etiqueta,
            descripcion=descripcion,
            esquema_json=schema_json,
            hash_esquema=hash_value,
            cambios_desde_anterior=json.dumps(cambios) if cambios else None,
            guardar_por=current_user.get("id"),
            activo=True,
        )
        
        # Mark previous as inactive
        if previous:
            previous.activo = False
        
        db.add(esquema)
        db.commit()
        
        return {
            "status": "success",
            "data": {
                "schema_id": esquema.id,
                "version": esquema.version,
                "hash": hash_value,
                "changes": cambios,
            }
        }
    
    except Exception as e:
        logger.error(f"Error saving schema: {e}")
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(e))


@router.get("/schemas/{db_name}/versions")
async def list_schema_versions(
    db_name: str,
    current_user: dict = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """List all saved schema versions for a database."""
    try:
        versions = db.query(EsquemaBaseDatos).filter(
            EsquemaBaseDatos.nombre_bd == db_name
        ).order_by(EsquemaBaseDatos.fecha_guardado.desc()).all()
        
        return {
            "status": "success",
            "data": [
                {
                    "id": v.id,
                    "version": v.version,
                    "etiqueta": v.etiqueta,
                    "descripcion": v.descripcion,
                    "hash": v.hash_esquema,
                    "fecha": v.fecha_guardado.isoformat(),
                    "activo": v.activo,
                }
                for v in versions
            ]
        }
    
    except Exception as e:
        logger.error(f"Error listing schema versions: {e}")
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(e))


@router.get("/schemas/{schema_id}/download")
async def download_schema_version(
    schema_id: int,
    current_user: dict = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Download a specific schema version."""
    try:
        esquema = db.query(EsquemaBaseDatos).filter(
            EsquemaBaseDatos.id == schema_id
        ).first()
        
        if not esquema:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Schema not found")
        
        return StreamingResponse(
            iter([esquema.esquema_json]),
            media_type="application/json",
            headers={"Content-Disposition": f"attachment; filename={esquema.nombre_bd}_v{esquema.version}.json"}
        )
    
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error downloading schema: {e}")
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(e))


def _compare_schemas(prev: dict, curr: dict) -> dict:
    """Compare two schema versions and report differences."""
    changes = {"tables_added": [], "tables_removed": [], "tables_modified": []}
    
    prev_tables = set(prev.get("tables", {}).keys())
    curr_tables = set(curr.get("tables", {}).keys())
    
    changes["tables_added"] = list(curr_tables - prev_tables)
    changes["tables_removed"] = list(prev_tables - curr_tables)
    
    # Check modified tables
    for table_name in prev_tables & curr_tables:
        prev_cols = {c["name"] for c in prev["tables"][table_name].get("columns", [])}
        curr_cols = {c["name"] for c in curr["tables"][table_name].get("columns", [])}
        
        if prev_cols != curr_cols:
            changes["tables_modified"].append({
                "table": table_name,
                "columns_added": list(curr_cols - prev_cols),
                "columns_removed": list(prev_cols - curr_cols),
            })
    
    return changes


# ===== GESTIÓN AVANZADA DE BACKUPS =====

@router.get("/backup/paths")
async def list_backup_paths(
    current_user: dict = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """List all configured backup paths."""
    try:
        manager = BackupManager(db)
        paths = manager.get_backup_paths()
        
        return {
            "status": "success",
            "data": paths,
        }
    
    except Exception as e:
        logger.error(f"Error listing backup paths: {e}")
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(e))


@router.post("/backup/paths")
async def add_backup_path(
    payload: dict,
    current_user: dict = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Add a new backup path."""
    try:
        path = str((payload or {}).get("path", "")).strip()
        is_active = bool((payload or {}).get("is_active", False))
        
        if not path:
            raise ValueError("Path is required")
        
        manager = BackupManager(db)
        success = manager.add_backup_path(path, is_active)
        
        if not success:
            raise ValueError("Failed to add backup path")
        
        return {
            "status": "success",
            "message": f"Backup path added: {path}",
            "data": manager.get_backup_paths(),
        }
    
    except Exception as e:
        logger.error(f"Error adding backup path: {e}")
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(e))


@router.put("/backup/paths/activate/{index}")
async def set_active_backup_path(
    index: int,
    current_user: dict = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Set the active backup path."""
    try:
        manager = BackupManager(db)
        paths = manager.get_backup_paths()
        
        if index < 0 or index >= len(paths):
            raise ValueError("Invalid path index")
        
        path = paths[index]["path"]
        success = manager.set_active_path(path)
        
        if not success:
            raise ValueError("Failed to set active path")
        
        return {
            "status": "success",
            "message": f"Active path set to: {path}",
            "data": manager.get_backup_paths(),
        }
    
    except Exception as e:
        logger.error(f"Error setting active backup path: {e}")
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(e))


@router.post("/backup/auto-config")
async def configure_auto_backup(
    payload: dict,
    current_user: dict = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Configure automatic backups."""
    try:
        enabled = bool((payload or {}).get("enabled", False))
        hour = int((payload or {}).get("hour", 2))
        retention_days = int((payload or {}).get("retention_days", 30))
        
        manager = BackupManager(db)
        
        if enabled:
            success = manager.enable_auto_backup(hour, retention_days)
        else:
            success = manager.disable_auto_backup()
        
        if not success:
            raise ValueError("Failed to configure auto-backup")
        
        return {
            "status": "success",
            "data": manager.get_auto_backup_config(),
        }
    
    except Exception as e:
        logger.error(f"Error configuring auto-backup: {e}")
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(e))


@router.get("/backup/auto-config")
async def get_auto_backup_config(
    current_user: dict = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Get auto-backup configuration."""
    try:
        manager = BackupManager(db)
        config = manager.get_auto_backup_config()
        
        return {
            "status": "success",
            "data": config,
        }
    
    except Exception as e:
        logger.error(f"Error getting auto-backup config: {e}")
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(e))


@router.post("/backup/cleanup")
async def cleanup_old_backups(
    payload: dict,
    current_user: dict = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Clean up backups older than N days."""
    try:
        days = int((payload or {}).get("days", 30))
        path = str((payload or {}).get("path", "")).strip() or None
        
        manager = BackupManager(db)
        result = manager.cleanup_old_backups(days, path)
        
        return {
            "status": "success",
            "data": result,
            "message": f"Deleted {result['deleted']} old backups",
        }
    
    except Exception as e:
        logger.error(f"Error cleaning up backups: {e}")
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(e))
