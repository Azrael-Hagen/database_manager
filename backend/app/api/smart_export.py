"""
Smart Export API – flexible data export with field selection, pattern filters,
and multiple output formats (CSV, Excel, TXT, DAT).

Security: every table and field name is validated against a strict whitelist
before being embedded in SQL.  Filter values are always passed as bound
parameters – there is no string interpolation of user-supplied values.
"""

from __future__ import annotations

import csv
import logging
import re
from io import BytesIO, StringIO
from typing import Any

from fastapi import APIRouter, Depends, HTTPException, status
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, field_validator
from sqlalchemy import inspect as sa_inspect, text
from sqlalchemy.orm import Session

from app.database.orm import get_db
from app.security import get_current_user, require_admin_role
from app.utils.export_formats import write_dat, write_txt

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api/smart-export", tags=["Exportación Inteligente"])

# ---------------------------------------------------------------------------
# Table access control
# ---------------------------------------------------------------------------

ALLOWED_TABLES_ANY_USER: frozenset[str] = frozenset(
    {
        "datos_importados",
        "lineas_telefonicas",
        "pagos_semanales",
        "agente_linea_asignaciones",
        "ladas_catalogo",
    }
)

ALLOWED_TABLES_ADMIN_ONLY: frozenset[str] = frozenset(
    {
        "usuarios",
        "auditoria_acciones",
        "import_logs",
    }
)

ALL_ALLOWED_TABLES: frozenset[str] = ALLOWED_TABLES_ANY_USER | ALLOWED_TABLES_ADMIN_ONLY

# ---------------------------------------------------------------------------
# Filter operators whitelist
# ---------------------------------------------------------------------------

ALLOWED_OPS: frozenset[str] = frozenset(
    {
        "eq",
        "neq",
        "contains",
        "starts_with",
        "ends_with",
        "gt",
        "lt",
        "gte",
        "lte",
        "is_null",
        "is_not_null",
        "in",
    }
)

_IDENTIFIER_RE = re.compile(r"^[a-zA-Z_][a-zA-Z0-9_]{0,63}$")


def _validate_identifier(name: str, label: str = "Identificador") -> str:
    """Raise ValueError if name is not a safe SQL identifier."""
    if not _IDENTIFIER_RE.match(name):
        raise ValueError(f"{label} inválido: '{name}'")
    return name


# ---------------------------------------------------------------------------
# Pydantic request schema
# ---------------------------------------------------------------------------


class FilterSpec(BaseModel):
    campo: str
    operador: str
    valor: str | None = None

    @field_validator("campo")
    @classmethod
    def validate_campo(cls, v: str) -> str:
        _validate_identifier(v, "Campo de filtro")
        return v

    @field_validator("operador")
    @classmethod
    def validate_operador(cls, v: str) -> str:
        if v not in ALLOWED_OPS:
            raise ValueError(f"Operador no permitido: '{v}'")
        return v


class ExportRequest(BaseModel):
    tabla: str
    campos: list[str]
    filtros: list[FilterSpec] = []
    formato: str = "csv"
    nombre_archivo: str | None = None
    limite: int | None = None

    @field_validator("tabla")
    @classmethod
    def validate_tabla(cls, v: str) -> str:
        if v not in ALL_ALLOWED_TABLES:
            raise ValueError(f"Tabla no permitida: '{v}'")
        return v

    @field_validator("formato")
    @classmethod
    def validate_formato(cls, v: str) -> str:
        if v not in {"csv", "excel", "txt", "dat"}:
            raise ValueError(f"Formato no soportado: '{v}'")
        return v

    @field_validator("campos")
    @classmethod
    def validate_campos(cls, v: list[str]) -> list[str]:
        if not v:
            raise ValueError("Debe especificar al menos un campo.")
        for c in v:
            _validate_identifier(c, "Campo")
        return v

    @field_validator("limite")
    @classmethod
    def validate_limite(cls, v: int | None) -> int | None:
        if v is not None and not (1 <= v <= 1_000_000):
            raise ValueError("límite debe estar entre 1 y 1 000 000.")
        return v


# ---------------------------------------------------------------------------
# WHERE clause builder (parameterized – no string interpolation of values)
# ---------------------------------------------------------------------------


def _build_where(filtros: list[FilterSpec]) -> tuple[str, dict]:
    """
    Return a parameterized WHERE clause string and a corresponding params dict.
    Column names are taken from the validated FilterSpec.campo (identifier-safe).
    All user-supplied *values* go into bind parameters only.
    """
    clauses: list[str] = []
    params: dict[str, Any] = {}

    for i, f in enumerate(filtros):
        col = f"`{f.campo}`"
        pname = f"p{i}"
        op = f.operador
        val = f.valor

        if op == "eq":
            clauses.append(f"{col} = :{pname}")
            params[pname] = val
        elif op == "neq":
            clauses.append(f"{col} != :{pname}")
            params[pname] = val
        elif op == "contains":
            clauses.append(f"{col} LIKE :{pname}")
            params[pname] = f"%{val}%"
        elif op == "starts_with":
            clauses.append(f"{col} LIKE :{pname}")
            params[pname] = f"{val}%"
        elif op == "ends_with":
            clauses.append(f"{col} LIKE :{pname}")
            params[pname] = f"%{val}"
        elif op == "gt":
            clauses.append(f"{col} > :{pname}")
            params[pname] = val
        elif op == "lt":
            clauses.append(f"{col} < :{pname}")
            params[pname] = val
        elif op == "gte":
            clauses.append(f"{col} >= :{pname}")
            params[pname] = val
        elif op == "lte":
            clauses.append(f"{col} <= :{pname}")
            params[pname] = val
        elif op == "is_null":
            clauses.append(f"{col} IS NULL")
        elif op == "is_not_null":
            clauses.append(f"{col} IS NOT NULL")
        elif op == "in":
            values = [s.strip() for s in (val or "").split(",") if s.strip()]
            if not values:
                continue
            placeholders = ", ".join(f":{pname}_{j}" for j in range(len(values)))
            clauses.append(f"{col} IN ({placeholders})")
            for j, vv in enumerate(values):
                params[f"{pname}_{j}"] = vv

    where = f"WHERE {' AND '.join(clauses)}" if clauses else ""
    return where, params


# ---------------------------------------------------------------------------
# Serialization helpers
# ---------------------------------------------------------------------------


def _ser(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, bool):
        return "1" if v else "0"
    return str(v)


def _to_csv(data: list[dict], campos: list[str]) -> bytes:
    output = StringIO()
    writer = csv.writer(output)
    writer.writerow(campos)
    for row in data:
        writer.writerow([_ser(row.get(c)) for c in campos])
    return output.getvalue().encode("utf-8-sig")


def _to_excel(data: list[dict], campos: list[str], sheet_name: str = "Datos") -> bytes:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]  # Excel sheet-name length limit

    header_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for col_idx, campo in enumerate(campos, 1):
        cell = ws.cell(row=1, column=col_idx, value=campo)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row_idx, row in enumerate(data, 2):
        for col_idx, campo in enumerate(campos, 1):
            ws.cell(row=row_idx, column=col_idx, value=_ser(row.get(campo)))

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


# ---------------------------------------------------------------------------
# Endpoints
# ---------------------------------------------------------------------------


@router.get("/tables")
async def list_export_tables(current_user: dict = Depends(get_current_user)):
    """
    Return the list of tables available for this user to export.
    Admin-only tables are included only for admin/super_admin roles.
    """
    is_admin = current_user.get("rol") in {"admin", "super_admin"}
    tables = list(ALLOWED_TABLES_ANY_USER)
    if is_admin:
        tables += list(ALLOWED_TABLES_ADMIN_ONLY)
    tables.sort()
    return {"status": "ok", "tablas": tables}


@router.get("/fields/{table_name}")
async def list_table_fields(
    table_name: str,
    current_user: dict = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """
    Return the column names (and types) for a given table.
    Uses SQLAlchemy inspect for DB-agnostic compatibility.
    """
    if table_name not in ALL_ALLOWED_TABLES:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Tabla no permitida: '{table_name}'",
        )

    is_admin = current_user.get("rol") in {"admin", "super_admin"}
    if table_name in ALLOWED_TABLES_ADMIN_ONLY and not is_admin:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Se requiere rol admin para acceder a esta tabla.",
        )

    try:
        inspector = sa_inspect(db.get_bind())
        raw_cols = inspector.get_columns(table_name)
        columns = [
            {
                "campo": col["name"],
                "tipo": str(col["type"]),
                "nullable": col.get("nullable", True),
            }
            for col in raw_cols
        ]
    except Exception as exc:
        logger.error("Error listing fields for %s: %s", table_name, exc)
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Tabla '{table_name}' no encontrada o inaccesible.",
        )

    return {"status": "ok", "tabla": table_name, "campos": columns}


@router.post("/export")
async def smart_export(
    request: ExportRequest,
    current_user: dict = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """
    Export rows from a table with:
    - Field selection (only the requested columns).
    - Pattern-based filters (eq, neq, contains, starts_with, ends_with, gt,
      lt, gte, lte, is_null, is_not_null, in).
    - Optional row limit.
    - Output formats: CSV, Excel (.xlsx), TXT (tab-delimited), DAT (pipe-delimited).

    All column and table names are whitelisted; filter values are always bound
    parameters – no SQL injection vector exists.

    Requires: any authenticated user for public tables; admin for restricted ones.
    """
    # Authorization check for admin-only tables
    is_admin = current_user.get("rol") in {"admin", "super_admin"}
    if request.tabla in ALLOWED_TABLES_ADMIN_ONLY and not is_admin:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Se requiere rol admin para exportar esta tabla.",
        )

    # Verify requested columns exist in the table
    try:
        inspector = sa_inspect(db.get_bind())
        all_cols = {col["name"] for col in inspector.get_columns(request.tabla)}
    except Exception:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Tabla '{request.tabla}' no encontrada.",
        )

    invalid_cols = [c for c in request.campos if c not in all_cols]
    if invalid_cols:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Campos no encontrados en '{request.tabla}': {invalid_cols}",
        )

    invalid_filter_cols = [f.campo for f in request.filtros if f.campo not in all_cols]
    if invalid_filter_cols:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Campos de filtro no encontrados: {invalid_filter_cols}",
        )

    # Build and execute query
    select_sql = ", ".join(f"`{c}`" for c in request.campos)
    where_sql, params = _build_where(request.filtros)
    limit_sql = f" LIMIT {request.limite}" if request.limite else ""
    query = f"SELECT {select_sql} FROM `{request.tabla}` {where_sql}{limit_sql}"

    try:
        result = db.execute(text(query), params)
        rows = result.fetchall()
    except Exception as exc:
        logger.error("smart-export query error: %s", exc)
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error ejecutando la consulta: {exc}",
        )

    data = [dict(zip(request.campos, row)) for row in rows]
    base_name = request.nombre_archivo or f"{request.tabla}_export"
    fmt = request.formato

    logger.info(
        "smart-export: user=%s tabla=%s campos=%d filas=%d formato=%s",
        current_user["username"],
        request.tabla,
        len(request.campos),
        len(data),
        fmt,
    )

    if fmt == "csv":
        body = _to_csv(data, request.campos)
        return StreamingResponse(
            iter([body]),
            media_type="text/csv; charset=utf-8",
            headers={"Content-Disposition": f"attachment; filename={base_name}.csv"},
        )

    if fmt == "excel":
        body = _to_excel(data, request.campos, base_name)
        return StreamingResponse(
            iter([body]),
            media_type=(
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            ),
            headers={
                "Content-Disposition": f"attachment; filename={base_name}.xlsx"
            },
        )

    if fmt == "txt":
        body = write_txt(data, request.campos)
        return StreamingResponse(
            iter([body]),
            media_type="text/plain; charset=utf-8",
            headers={
                "Content-Disposition": f"attachment; filename={base_name}.txt"
            },
        )

    # fmt == "dat"
    body = write_dat(data, request.campos)
    return StreamingResponse(
        iter([body]),
        media_type="application/octet-stream",
        headers={"Content-Disposition": f"attachment; filename={base_name}.dat"},
    )
