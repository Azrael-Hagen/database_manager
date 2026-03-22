"""Utility functions for exporting database data to CSV and Excel formats."""

import csv
import json
import logging
from io import BytesIO, StringIO
from datetime import datetime
from typing import List, Dict, Optional, Any

from sqlalchemy.orm import Session
from sqlalchemy import inspect, text

logger = logging.getLogger(__name__)


def _serialize_value(value: Any) -> str:
    """Convert database values to CSV-safe strings."""
    if value is None:
        return ""
    if isinstance(value, (datetime,)):
        return value.isoformat()
    if isinstance(value, bool):
        return "Sí" if value else "No"
    if isinstance(value, dict):
        return json.dumps(value, ensure_ascii=False)
    return str(value)


def export_table_to_csv(db: Session, db_name: str, table_name: str, limit: Optional[int] = None) -> str:
    """
    Export table data to CSV format.
    
    Args:
        db: Database session
        db_name: Database name
        table_name: Table name
        limit: Optional row limit
    
    Returns:
        CSV string content
    """
    try:
        # Change to target database
        db.execute(text(f"USE `{db_name}`"))
        
        # Get table columns
        result = db.execute(text(f"DESCRIBE `{table_name}`"))
        columns = [row[0] for row in result.fetchall()]
        
        if not columns:
            raise ValueError(f"No columns found in {table_name}")
        
        # Build query
        limit_clause = f" LIMIT {limit}" if limit else ""
        query = f"SELECT {', '.join(f'`{col}`' for col in columns)} FROM `{table_name}`{limit_clause}"
        
        result = db.execute(text(query))
        rows = result.fetchall()
        
        # Build CSV
        output = StringIO()
        writer = csv.writer(output)
        
        # Write header
        writer.writerow(columns)
        
        # Write rows
        for row in rows:
            writer.writerow([_serialize_value(val) for val in row])
        
        return output.getvalue()
    
    except Exception as e:
        logger.error(f"Error exporting {table_name} to CSV: {e}")
        raise ValueError(f"Error exporting table: {str(e)}")


def export_datos_importados_to_csv(db: Session, with_pagos: bool = False) -> str:
    """
    Export agentes (datos_importados) with optional payments.
    
    Args:
        db: Database session
        with_pagos: Include payment information
    
    Returns:
        CSV string content
    """
    try:
        from app.models import DatoImportado, PagoSemanal
        
        agentes = db.query(DatoImportado).filter(DatoImportado.es_activo.is_(True)).all()
        
        output = StringIO()
        
        # Determine headers
        headers = [
            "ID", "UUID", "Nombre", "Email", "Teléfono", "Empresa", 
            "Ciudad", "País", "Alias", "Ubicación", "FP", "FC", "Grupo",
            "Número VoIP", "Créado por", "Fecha Creación"
        ]
        
        if with_pagos:
            headers.extend(["Semana", "Pagado", "Monto", "Fecha Pago"])
        
        writer = csv.writer(output)
        writer.writerow(headers)
        
        # Write rows
        for agente in agentes:
            datos_add = json.loads(agente.datos_adicionales) if agente.datos_adicionales else {}
            
            row = [
                agente.id,
                agente.uuid,
                agente.nombre,
                agente.email,
                agente.telefono,
                agente.empresa,
                agente.ciudad,
                agente.pais,
                datos_add.get("alias", ""),
                datos_add.get("ubicacion", ""),
                datos_add.get("fp", ""),
                datos_add.get("fc", ""),
                datos_add.get("grupo", ""),
                datos_add.get("numero_voip", ""),
                agente.creado_por,
                agente.fecha_creacion.isoformat(),
            ]
            
            if with_pagos:
                # Get latest payment for this agent
                pago = db.query(PagoSemanal).filter(
                    PagoSemanal.agente_id == agente.id
                ).order_by(PagoSemanal.semana_inicio.desc()).first()
                
                if pago:
                    row.extend([
                        pago.semana_inicio.isoformat(),
                        _serialize_value(pago.pagado),
                        pago.monto,
                        pago.fecha_pago.isoformat() if pago.fecha_pago else "",
                    ])
                else:
                    row.extend(["", "No", "0", ""])
            
            writer.writerow(row)
        
        return output.getvalue()
    
    except Exception as e:
        logger.error(f"Error exporting agentes to CSV: {e}")
        raise ValueError(f"Error exporting agentes: {str(e)}")


def export_to_excel(db: Session, db_name: str, table_name: str, limit: Optional[int] = None) -> bytes:
    """
    Export table data to Excel format.
    
    Args:
        db: Database session
        db_name: Database name
        table_name: Table name
        limit: Optional row limit
    
    Returns:
        Excel file bytes
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        
        # Get CSV data first
        csv_data = export_table_to_csv(db, db_name, table_name, limit)
        
        # Parse CSV
        reader = csv.reader(StringIO(csv_data))
        rows = list(reader)
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = table_name[:31]  # Excel has 31-char limit
        
        # Style header
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        # Add rows
        for row_idx, row in enumerate(rows, 1):
            for col_idx, cell_value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=cell_value)
                
                if row_idx == 1:  # Header row
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center")
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to bytes
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output.getvalue()
    
    except ImportError:
        logger.error("openpyxl not installed. Install with: pip install openpyxl")
        raise ValueError("Excel export requires openpyxl package")
    except Exception as e:
        logger.error(f"Error exporting to Excel: {e}")
        raise ValueError(f"Error exporting to Excel: {str(e)}")


def export_schema_to_json(db: Session, db_name: str) -> str:
    """
    Export database schema as JSON documentation.
    
    Args:
        db: Database session
        db_name: Database name
    
    Returns:
        JSON string with schema info
    """
    try:
        db.execute(text(f"USE `{db_name}`"))
        
        # Get all tables
        result = db.execute(text("SHOW TABLES"))
        table_names = [row[0] for row in result.fetchall()]
        
        schema = {
            "database": db_name,
            "exported": datetime.utcnow().isoformat(),
            "tables": {}
        }
        
        for table_name in table_names:
            # Get columns info
            result = db.execute(text(f"DESCRIBE `{table_name}`"))
            columns = []
            for row in result.fetchall():
                columns.append({
                    "name": row[0],
                    "type": row[1],
                    "null": row[2] == "YES",
                    "key": row[3] or None,
                    "default": row[4],
                    "extra": row[5] or None,
                })
            
            # Get table creation SQL
            result = db.execute(text(f"SHOW CREATE TABLE `{table_name}`"))
            create_sql = result.fetchone()[1]
            
            schema["tables"][table_name] = {
                "columns": columns,
                "create_sql": create_sql,
            }
        
        return json.dumps(schema, ensure_ascii=False, indent=2)
    
    except Exception as e:
        logger.error(f"Error exporting schema: {e}")
        raise ValueError(f"Error exporting schema: {str(e)}")
